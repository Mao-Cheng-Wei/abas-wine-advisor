import os
import json
import re
import html
import requests
from pathlib import Path

import streamlit as st
import tiktoken
from openpyxl import load_workbook
from openai import OpenAI
from pydantic import BaseModel, Field
from llama_index.core import SimpleDirectoryReader, VectorStoreIndex, Settings
from llama_index.embeddings.fastembed import FastEmbedEmbedding


GEMINI_MODELS = [
    ("gemini-3.7-flash", "Gemini 3.7 Flash", "low"),
    ("gemini-3.6-flash", "Gemini 3.6 Flash", "minimal"),
]

APP_URL = "https://abas-wine-advisor-2ju5inrreherxujphsnu7y.streamlit.app"
CONTACT_PAGE = "https://dagc.com.tw/contact"
DEFAULT_LINE_URL = "https://lin.ee/Hw3jvZI"
PRODUCT_IMAGE_DIR = Path("images/products")


class AdvisorRecommendation(BaseModel):
    name: str = Field(description="推薦酒款名稱，必須完全沿用提供的產品名稱")
    reason: str = Field(description="2到3句繁體中文推薦理由")


class AdvisorResult(BaseModel):
    personality: str = Field(description="80到120字、2到3句的繁體中文風味人格輪廓")
    recommendations: list[AdvisorRecommendation] = Field(description="依照既定排名逐一提供推薦酒款與理由")


st.set_page_config(
    page_title="安貝斯風味人格選酒顧問",
    page_icon="🍷",
    layout="centered",
)

st.markdown(
    """
    <style>
    .block-container {padding-top:4.6rem!important;padding-bottom:3rem!important;max-width:980px;}
    .hero-kicker{font-size:14px;letter-spacing:.12em;color:#A9782A;font-weight:800;margin-bottom:6px;}
    .hero-title{font-size:38px;font-weight:800;line-height:1.15;color:#2F3140;margin:0 0 8px;}
    .hero-sub{font-size:18px;color:#666;line-height:1.75;margin-bottom:12px;}
    .section-title{font-size:30px;font-weight:800;color:#2F3140;margin:26px 0 12px;}
    .chip{display:inline-block;background:#F7F0E4;color:#7A5520;padding:7px 14px;border-radius:999px;margin:4px;font-size:16px;border:1px solid #E8D6B8;}
    .mini-note{font-size:14px;color:#888;line-height:1.6;}
    .ai-main-title{text-align:center;font-size:38px;font-weight:800;color:#2F3140;margin-top:28px;margin-bottom:6px;}
    .ai-subtitle{text-align:center;font-size:16px;color:#7A7A7A;margin-bottom:18px;}
    .brand-line{height:1px;background:#E8DFCF;margin:18px 0 22px;}
    @media (max-width:640px){
      .block-container{padding-top:3.6rem!important;padding-left:1rem!important;padding-right:1rem!important;}
      .hero-title{font-size:29px}.hero-sub{font-size:16px}.section-title{font-size:25px}.ai-main-title{font-size:31px}
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_secret(name, default=None):
    try:
        value = st.secrets[name]
        return value if value else default
    except Exception:
        return os.getenv(name, default)


@st.cache_data(ttl=86400, show_spinner=False)
def get_official_product_image(product_url):
    if not product_url or "dagc.com.tw" not in str(product_url):
        return None
    try:
        r = requests.get(
            str(product_url),
            timeout=8,
            headers={"User-Agent": "Mozilla/5.0 ABAS-Wine-Advisor/1.0"},
        )
        r.raise_for_status()
        page = r.text
        patterns = [
            r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)',
            r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']',
            r'https://static\.wixstatic\.com/media/[^"\'<> ]+',
        ]
        for pattern in patterns:
            m = re.search(pattern, page, flags=re.I)
            if m:
                value = m.group(1) if m.lastindex else m.group(0)
                return html.unescape(value).replace('\\u002F', '/')
    except Exception:
        return None
    return None


def find_product_image(product_id, product_url=None):
    if PRODUCT_IMAGE_DIR.exists():
        for ext in ("jpg", "jpeg", "png", "webp"):
            p = PRODUCT_IMAGE_DIR / f"{product_id}.{ext}"
            if p.exists():
                return str(p)
    return get_official_product_image(product_url)


hero_left, hero_right = st.columns([1, 5], vertical_alignment="center")
with hero_left:
    st.image("images/ABAS_logo.jpg", width=86)
with hero_right:
    st.markdown('<div class="hero-kicker">ABAS FLAVOR PERSONALITY</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-title">安貝斯風味人格選酒顧問</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-sub">先不談酒名，從生活、香氣與口感直覺出發，找出更貼近你的風味方向。</div>', unsafe_allow_html=True)

st.markdown('<div class="brand-line"></div>', unsafe_allow_html=True)
st.caption("台中大安風土｜純糧固態發酵蒸餾｜30年釀酒傳承")

questions = [
    {"id":"Q01","title":"社交能量","question":"今晚最想把時間留給哪一種狀態？","options":{"A":{"text":"一個人安靜整理思緒","tags":["安靜","內省","慢飲"]},"B":{"text":"和一位熟悉的人好好聊天","tags":["親密","柔和","分享"]},"C":{"text":"三五好友輕鬆聚會","tags":["活潑","分享","易飲"]},"D":{"text":"把氣氛推到最高點","tags":["強烈","慶祝","高能量"]}}},
    {"id":"Q02","title":"生活節奏","question":"你的理想週末比較像哪一幕？","options":{"A":{"text":"晨光、書、沒有行程","tags":["清爽","細緻","低刺激"]},"B":{"text":"午後茶席與一段長談","tags":["茶香","雅緻","柔順"]},"C":{"text":"黃昏市集與街邊小吃","tags":["果香","鮮明","親切"]},"D":{"text":"夜晚派對與即興冒險","tags":["濃郁","強烈","高酒感"]}}},
    {"id":"Q03","title":"自然意象","question":"哪一種風景最像你現在的心情？","options":{"A":{"text":"竹林與薄霧","tags":["清幽","草本","內斂"]},"B":{"text":"金色夕陽與柑橘園","tags":["柑橘","明亮","溫暖"]},"C":{"text":"海風、礁石與鹹味空氣","tags":["海風","鹹感","俐落"]},"D":{"text":"木屋、火光與乾燥木香","tags":["木質","煙燻","成熟"]}}},
    {"id":"Q04","title":"味覺性格","question":"吃甜點時，你通常偏向？","options":{"A":{"text":"幾乎不甜，重視原味","tags":["乾型","穀物","純粹"]},"B":{"text":"微甜即可，尾韻乾淨","tags":["低甜","清爽","平衡"]},"C":{"text":"酸甜平衡，容易入口","tags":["酸甜","果香","易飲"]},"D":{"text":"濃郁甜香，層次越多越好","tags":["甜潤","濃郁","桶陳"]}}},
    {"id":"Q05","title":"口感偏好","question":"你喜歡的飲品口感比較接近？","options":{"A":{"text":"清脆俐落，像冰涼氣泡水","tags":["清爽","輕盈","俐落"]},"B":{"text":"滑順柔和，沒有壓迫感","tags":["柔順","圓潤","入門"]},"C":{"text":"厚實飽滿，入口有存在感","tags":["厚實","醇厚","熟成"]},"D":{"text":"強勁集中，能感覺到力量","tags":["高酒感","強勁","原漿"]}}},
    {"id":"Q06","title":"香氣直覺","question":"不看酒名，你最想靠近哪一種香氣？","options":{"A":{"text":"梅子、蜜桃與果乾","tags":["梅果","蜜桃","酸甜"]},"B":{"text":"柑橘、鳳梨與新鮮果皮","tags":["柑橘","熱帶水果","明亮"]},"C":{"text":"茶葉、香草與草本植物","tags":["茶香","草本","清新"]},"D":{"text":"烘烤穀物、木桶與太妃糖","tags":["穀香","木桶","焦糖"]}}},
    {"id":"Q07","title":"個性表達","question":"朋友通常怎麼形容你？","options":{"A":{"text":"安靜但有自己的深度","tags":["內斂","細緻","熟成"]},"B":{"text":"溫和、可靠、容易親近","tags":["柔順","平衡","大眾接受"]},"C":{"text":"有趣、有創意、常帶來驚喜","tags":["特色","草本","創新"]},"D":{"text":"果斷、有主見、氣場明確","tags":["強烈","高酒感","收藏"]}}},
    {"id":"Q08","title":"冒險程度","question":"面對沒喝過的風味，你會？","options":{"A":{"text":"先選經典安全款","tags":["經典","入門","平衡"]},"B":{"text":"熟悉中有一點變化最好","tags":["親切","特色","低風險"]},"C":{"text":"願意嘗試地方植物或特殊香氣","tags":["創新","在地","草本"]},"D":{"text":"越少見、越有挑戰越想試","tags":["限量","高酒感","實驗性"]}}},
    {"id":"Q09","title":"飲用情境","question":"你希望這瓶酒主要出現在？","options":{"A":{"text":"獨處慢飲，讓自己沉澱","tags":["慢飲","深度","熟成"]},"B":{"text":"餐桌上搭配料理","tags":["搭餐","平衡","乾淨"]},"C":{"text":"朋友聚會或調酒時刻","tags":["調酒","活潑","分享"]},"D":{"text":"重要節日或值得收藏的時刻","tags":["贈禮","收藏","尊榮"]}}},
    {"id":"Q10","title":"酒精感接受度","question":"你期待酒精帶來多少存在感？","options":{"A":{"text":"越低越好，以果香為主","tags":["低酒精","果香","易飲"]},"B":{"text":"有感但柔和，不想太刺激","tags":["中低酒感","柔順","平衡"]},"C":{"text":"明顯、有溫度，但仍要有層次","tags":["中高酒感","醇厚","熟成"]},"D":{"text":"喜歡強勁、集中、帶衝擊力","tags":["高酒感","原漿","強勁"]}}},
]


def get_alcohol_level(alcohol_text):
    if alcohol_text is None:
        return "待確認"
    text = str(alcohol_text).replace("%", "").replace("vol", "").strip()
    try:
        value = float(text)
        if value < 20: return "低"
        if value < 50: return "中高"
        return "高"
    except Exception:
        return "待確認"


def split_tags(tag_text):
    if tag_text is None: return []
    text = str(tag_text).strip()
    if text in ["", "待補", "待確認", "None"]: return []
    for mark in ["；", ";", ",", "，"]: text = text.replace(mark, "、")
    return [t.strip() for t in text.split("、") if t.strip()]


@st.cache_data
def load_products_from_excel():
    wb = load_workbook("data/安貝斯產品知識庫_維護主檔_補標籤.xlsx", data_only=True)
    ws = wb["知識庫產品主檔"]
    headers = [c.value for c in ws[1]]
    hm = {h:i for i,h in enumerate(headers) if h}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[hm["產品編號"]]: continue
        if str(row[hm["產品類型"]]).strip() != "酒類": continue
        current,status,alcohol = row[hm["是否現行商品"]],row[hm["商品狀態"]],row[hm["酒精度"]]
        out.append({
            "id":row[hm["產品編號"]],"name":row[hm["產品名稱"]],"tags":split_tags(row[hm["風味標籤"]]),
            "alcohol_level":get_alcohol_level(alcohol),"status":status,
            "recommendable":str(current).strip()=="是" and str(status).strip()=="販售中",
            "url":row[hm["來源網址"]],"price":row[hm["價格"]],"alcohol":alcohol,
        })
    return out


products = load_products_from_excel()


@st.cache_resource(show_spinner="第一次啟動：正在載入中文 Embedding 模型與建立 RAG 索引...")
def build_rag_index():
    Settings.embed_model = FastEmbedEmbedding(model_name="BAAI/bge-small-zh-v1.5")
    Settings.tokenizer = tiktoken.get_encoding("cl100k_base").encode
    docs = SimpleDirectoryReader(input_dir="data", recursive=True, required_exts=[".md", ".txt"]).load_data()
    return VectorStoreIndex.from_documents(docs)


def get_gemini_client():
    key = get_secret("GEMINI_API_KEY")
    if not key: raise RuntimeError("尚未設定 GEMINI_API_KEY。")
    return OpenAI(base_url="https://generativelanguage.googleapis.com/v1beta/openai/", api_key=key)


def call_gemini_structured_with_fallback(client, messages):
    errors=[]
    for model_slug,model_label,effort in GEMINI_MODELS:
        try:
            completion=client.beta.chat.completions.parse(
                model=model_slug,
                messages=messages,
                temperature=0.1,
                reasoning_effort=effort,
                max_tokens=2200,
                timeout=120,
                response_format=AdvisorResult,
            )
            parsed=completion.choices[0].message.parsed
            if parsed is None: raise RuntimeError("模型沒有回傳可解析的結構化結果")
            return parsed,model_label
        except Exception as e:
            errors.append(f"{model_label}: {e}")
    raise RuntimeError("Gemini 雲端模型目前無法完成分析。"+" | ".join(errors))


def validate_advisor_payload(payload,top3):
    if not payload.personality.strip() or len(payload.recommendations)!=len(top3): return False
    for idx,product in enumerate(top3):
        item=payload.recommendations[idx]
        if item.name.strip()!=str(product["name"]).strip() or not item.reason.strip(): return False
    return True


def build_advisor_messages(rag_context,user_scores,top3,retry_note=""):
    rec=[{"rank":idx,"name":p["name"],"alcohol_level":p["alcohol_level"],"matched_tags":p["matched_tags"]} for idx,p in enumerate(top3,1)]
    prompt=f"""你是安貝斯風味人格選酒顧問。推薦酒款與順序已固定，你只負責解讀。
【安貝斯知識】
{rag_context}
【使用者風味標籤】
{json.dumps(user_scores,ensure_ascii=False)}
【固定推薦結果】
{json.dumps(rec,ensure_ascii=False)}
請依結構化欄位輸出：personality 80～120字、2～3句；每款 reason 2～3句。
酒名、筆數與順序必須完全一致；理由只能根據提供的知識與標籤。
不得自行增加年份、日期、原料、產地、獎項、庫存、價格；資料不足直接略過。
不要提 RAG、資料庫、模型、規則引擎、分數或技術流程。使用繁體中文，三款理由不要重複。
{retry_note}"""
    return [
        {"role":"system","content":"你是安貝斯品牌的繁體中文選酒顧問。只根據提供資料回答，不可臆測。"},
        {"role":"user","content":prompt},
    ]


def generate_advisor_result(client,rag_context,user_scores,top3):
    last_error=None
    for attempt in range(2):
        retry_note="" if attempt==0 else "上一次內容未通過驗證，請補齊全部酒款且保持名稱與順序。"
        try:
            payload,used_model=call_gemini_structured_with_fallback(client,build_advisor_messages(rag_context,user_scores,top3,retry_note))
            if validate_advisor_payload(payload,top3): return payload,used_model
            last_error=RuntimeError("AI 回覆欄位或酒款順序不完整")
        except Exception as e:
            last_error=e
    raise RuntimeError(f"AI 解讀結構驗證失敗：{last_error}")


if "step" not in st.session_state: st.session_state.step=0
if "scores" not in st.session_state: st.session_state.scores={}
if "finished" not in st.session_state: st.session_state.finished=False

if not st.session_state.finished:
    st.progress((st.session_state.step+1)/len(questions))
    st.caption(f"第 {st.session_state.step+1} / {len(questions)} 題")
    q=questions[st.session_state.step]
    st.subheader(f"{q['id']}｜{q['title']}")
    st.write(q["question"])
    answer=st.radio("請選擇：",list(q["options"].keys()),format_func=lambda x:f"{x}. {q['options'][x]['text']}",key=f"q_{st.session_state.step}")
    if st.button("下一題",use_container_width=True,type="primary"):
        for tag in q["options"][answer]["tags"]: st.session_state.scores[tag]=st.session_state.scores.get(tag,0)+3
        if st.session_state.step+1>=len(questions): st.session_state.finished=True
        else: st.session_state.step+=1
        st.rerun()
    st.markdown('<div class="mini-note">本測驗提供風味探索與產品認識，不代表飲酒必要性。未滿 18 歲請勿飲酒，飲酒勿駕車。</div>',unsafe_allow_html=True)
else:
    user_scores=st.session_state.scores
    if "低酒精" in user_scores or "中低酒感" in user_scores: ual="低"
    elif "中高酒感" in user_scores: ual="中高"
    elif "高酒感" in user_scores: ual="高"
    else: ual=None

    results=[]
    for p in products:
        if not p["recommendable"]: continue
        if ual=="低" and p["alcohol_level"] in ["中高","高"]: continue
        if ual=="中高" and p["alcohol_level"]=="高": continue
        score=0; matched=[]
        for tag in p["tags"]:
            if tag in user_scores: score+=3; matched.append(tag)
        if p["alcohol_level"]==ual: score+=4
        if p["status"]=="販售中": score+=2
        results.append({**p,"score":score,"matched_tags":matched})
    results.sort(key=lambda x:x["score"],reverse=True)
    top3=results[:3]
    top_tags=sorted(user_scores.items(),key=lambda x:x[1],reverse=True)[:8]

    st.markdown('<div class="section-title">你的風味人格</div>',unsafe_allow_html=True)
    st.markdown("".join(f'<span class="chip">{tag}</span>' for tag,_ in top_tags),unsafe_allow_html=True)
    st.markdown('<div class="section-title">為你挑出的酒款</div>',unsafe_allow_html=True)
    labels=["最像你的酒","另一種可能","想挑戰的酒"]
    for i,p in enumerate(top3):
        with st.container(border=True):
            img_col,info_col=st.columns([1.1,2.3],vertical_alignment="center")
            with img_col:
                image_path=find_product_image(p["id"], p.get("url"))
                if image_path: st.image(image_path,use_container_width=True)
                else: st.image("images/ABAS_logo.jpg",width=105); st.caption("官方商品圖載入中／待補")
            with info_col:
                st.caption(labels[i]); st.subheader(p["name"])
                if p["matched_tags"]: st.write("風味契合："+"、".join(p["matched_tags"]))
                if p["alcohol"]: st.write(f"酒精度：{p['alcohol']}")
                if p["price"]: st.write(f"參考價格：NT${p['price']}")
                if p["url"]: st.link_button("查看官方產品",p["url"],use_container_width=True)

    if top3:
        try:
            index=build_rag_index()
            retriever=index.as_retriever(similarity_top_k=3)
            rq=f"使用者風味標籤：{user_scores}\n推薦酒款：{[p['name'] for p in top3]}\n請找與這些風味人格、產品特色與推薦規則最相關的安貝斯知識。"
            nodes=retriever.retrieve(rq)
            rag_context="\n\n".join(n.get_content() for n in nodes)[:8000]
            client=get_gemini_client()
            with st.spinner("正在整理你的專屬風味解讀..."):
                advisor,used_model=generate_advisor_result(client,rag_context,user_scores,top3)
            st.markdown('<div class="ai-main-title">安貝斯 AI 顧問解讀</div>',unsafe_allow_html=True)
            st.markdown('<div class="ai-subtitle">從你的風味偏好出發，看看哪一款最貼近現在的你</div>',unsafe_allow_html=True)
            with st.container(border=True):
                st.markdown("### 你的風味人格輪廓"); st.write(advisor.personality)
                for idx,item in enumerate(advisor.recommendations):
                    st.markdown(f"### {labels[idx]}"); st.markdown(f"**{item.name}**"); st.write(item.reason)
            with st.expander("技術資訊"): st.caption(f"本次雲端模型：{used_model}")
        except Exception as e:
            st.info("AI 顧問暫時忙碌中；上方推薦結果仍可正常使用。")
            with st.expander("技術資訊"): st.caption(str(e))

    st.markdown('<div class="section-title">想進一步了解這款酒？</div>',unsafe_allow_html=True)
    with st.container(border=True):
        st.subheader("讓安貝斯接著陪你選")
        st.write("如果你想依送禮、聚會、搭餐或個人口味再挑得更精準，可以直接與安貝斯聯絡。")
        st.markdown("""**安貝斯聯絡資訊**  
台中市大安區頂安里中山北路330號  
聯絡電話：04-26886059 / 04-2688-8318  
E-mail：daanabas1989@gmail.com  
官方 LINE：@abas  
服務時間：週一～週五 9:00–18:00""")
        line_url=get_secret("LINE_CONTACT_URL",DEFAULT_LINE_URL); contact_url=get_secret("CONTACT_URL",CONTACT_PAGE)
        c1,c2=st.columns(2)
        with c1: st.link_button("官方 LINE｜@abas",line_url,use_container_width=True)
        with c2: st.link_button("聯絡安貝斯",contact_url,use_container_width=True)
        c3,c4=st.columns(2)
        with c3: st.link_button("撥打 04-26886059","tel:0426886059",use_container_width=True)
        with c4: st.link_button("Email 安貝斯","mailto:daanabas1989@gmail.com",use_container_width=True)

    st.markdown('<div class="section-title">分享我的風味結果</div>',unsafe_allow_html=True)
    share_lines=["我剛完成安貝斯風味人格選酒測驗 🍷","我的風味關鍵字："+"、".join(tag for tag,_ in top_tags[:5])]
    if top3:
        share_lines.append("最像我的酒："+top3[0]["name"])
        if len(top3)>1: share_lines.append("另一種可能："+top3[1]["name"])
    share_lines.append("也來測測看："+APP_URL)
    share_text="\n".join(share_lines)
    st.caption("點右上角複製圖示，即可貼到 LINE、Facebook 或訊息中。")
    st.code(share_text,language=None)
    st.download_button("下載我的風味結果",data=share_text,file_name="ABAS_風味人格結果.txt",mime="text/plain",use_container_width=True)
    st.caption("未滿 18 歲請勿飲酒｜飲酒勿駕車｜本測驗為風味探索與產品認識用途")
    if st.button("重新測驗",use_container_width=True):
        st.session_state.step=0; st.session_state.scores={}; st.session_state.finished=False; st.rerun()
